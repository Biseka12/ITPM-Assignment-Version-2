import time
import argparse
import openpyxl
from playwright.sync_api import sync_playwright


def is_sinhala(text):
    return any('\u0D80' <= c <= '\u0DFF' for c in text)


def run_test():
    parser = argparse.ArgumentParser()

    parser.add_argument('--excel', required=True)
    parser.add_argument('--url', required=True)
    parser.add_argument('--wait-ms', type=int, default=5000)
    parser.add_argument('--type-delay-ms', type=int, default=50)
    parser.add_argument('--slow-mo-ms', type=int, default=200)
    parser.add_argument('--save-every', type=int, default=1)
    parser.add_argument('--keep-open', action='store_true')

    args = parser.parse_args()

    wb = openpyxl.load_workbook(args.excel)
    sheet = wb.active

    headers = [cell.value for cell in sheet[1]]

    tc_col = input_col = expected_col = actual_col = status_col = None

    for i, h in enumerate(headers, start=1):
        if h and "TC" in str(h):
            tc_col = i
        elif h and "Input" in str(h):
            input_col = i
        elif h and "Expected" in str(h):
            expected_col = i
        elif h and "Actual" in str(h):
            actual_col = i
        elif h and "Status" in str(h):
            status_col = i

    print(f"Columns found: TC={tc_col}, Input={input_col}, Expected={expected_col}")

    with sync_playwright() as p:

        browser = p.chromium.launch(
            headless=False,
            slow_mo=args.slow_mo_ms
        )

        page = browser.new_page()

        print(f"Loading: {args.url}")
        page.goto(args.url)

        page.wait_for_timeout(args.wait_ms)
        page.wait_for_load_state("networkidle")

        print("Frontend loaded")

        # -------------------------
        # FIND INPUT BOX
        # -------------------------
        input_box = None

        selectors = [
            "textarea",
            "input[type='text']",
            "div[contenteditable='true']",
            "div[role='textbox']"
        ]

        for sel in selectors:
            try:
                loc = page.locator(sel).first
                if loc.is_visible():
                    input_box = loc
                    print(f"Input found: {sel}")
                    break
            except:
                pass

        if not input_box:
            print("ERROR: Input not found")
            page.screenshot(path="error.png")
            browser.close()
            return

        total_rows = sheet.max_row
        print(f"Running {total_rows - 1} test cases")

        # -------------------------
        # TEST LOOP
        # -------------------------
        for row in range(2, total_rows + 1):

            tc = sheet.cell(row=row, column=tc_col).value
            inp = sheet.cell(row=row, column=input_col).value
            expected = sheet.cell(row=row, column=expected_col).value

            if not inp:
                print(f"Row {row}: empty input")
                continue

            print(f"\nRunning {tc}")

            try:
                # -------------------------
                # INPUT (FIXED)
                # -------------------------
                input_box.click()
                input_box.fill("")
                input_box.fill(inp)
                input_box.press("Enter")

                # wait for response
                page.wait_for_timeout(7000)

                # -------------------------
                # FIXED OUTPUT CAPTURE (IMPORTANT FIX)
                # -------------------------

                texts = page.locator("div").all_inner_texts()

                output = ""

                for t in reversed(texts):
                    if not t:
                        continue

                    t = t.strip()

                    # ❌ ignore UI noise
                    if "PixelsSuite" in t:
                        continue
                    if "©" in t:
                        continue
                    if "Output (Sinhala)" in t:
                        continue
                    if inp in t:
                        continue
                    if len(t) < 2:
                        continue

                    # ✔ must be Sinhala or meaningful text
                    if is_sinhala(t):
                        output = t
                        break

                print("OUTPUT:", output if output else "[EMPTY]")

                # -------------------------
                # STATUS LOGIC (FIXED)
                # -------------------------
                if output and expected and str(expected).strip() in output:
                    status = "PASS"
                elif output:
                    status = "FAIL (mismatch)"
                else:
                    status = "FAIL (no output)"

                sheet.cell(row=row, column=actual_col, value=output)
                sheet.cell(row=row, column=status_col, value=status)

                print("STATUS:", status)

            except Exception as e:
                print("ERROR:", e)
                sheet.cell(row=row, column=actual_col, value=str(e))
                sheet.cell(row=row, column=status_col, value="ERROR")

            if row % args.save_every == 0:
                wb.save(args.excel)
                print("Saved progress")

        wb.save(args.excel)
        print("\nDONE - Results saved")

        if args.keep_open:
            input("Press Enter to close browser...")

        browser.close()


if __name__ == "__main__":
    run_test()